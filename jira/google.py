import jira.client
from jira.client import JIRA
from openpyxl import Workbook

#jira_options={'server': 'http://example.com'}
#jira=JIRA(options=jira_options,basic_auth=('admin','password'))
#credentials={"serviceAccount" : "farid.roshan@altimetrik.com", "serviceAccountToken" : "cA1AvOAgGdMaccejfgwy8919"}
credentials={"serviceAccount" : "farid.roshan@altimetrik.com", "serviceAccountToken" : "cA1AvOAgGdMaccejfgwy8919"}
jiraURL= "https://altimetrik.atlassian.net"
jira= requests.get(jiraURL, credentials)
#jiraURL= {'server':"https://altimetrik.atlassian.net"}
#jira=JIRA(options=jiraURL, basic_auth=("serviceAcount","serviceAccountToken"))

key_list = []
summary_list = []

#Add additional lists for fields here
#Example: 
#description_list = []



issues_in_project = JIRA.search_issues('project=VCQI AND assignee= DEAN')

for issue in issues_in_project:
    key_list.append(issue.key)
    summary_list.append(issue.fields.summary)

# Add additional fields returned here
# Example:
#. description_list.append(issue.fields.description)


wb = Workbook()
ws = wb.active
key_row = 1
summary_row = 1
#add additional "$FIELD_row = 1" entries here so the field results start at row 1
#Example:
#description_row = 1


start_column = 1



for key in key_list:
 ws.cell(row=key_row, column=start_column).value = key
 key_row += 1

for summary in summary_list:
 ws.cell(row=summary_row, column=start_column+1).value = summary
 summary_row += 1

# add additional fields here 
#Example:
#for description in description_list:
 # ws.cell(row=description_row, column=start_column+2).value = description
 # description_row += 1
